import pandas as pd
import os
from openpyxl import load_workbook


def process_data(archivo_excel):
    # Crear directorio de data
    os.makedirs("data",exist_ok=True)

    # 1. Crear camiones.csv
    df_camiones = pd.read_excel(archivo_excel, sheet_name="Pendientes por recibir detalle")
    df_camiones = df_camiones[['CARGA', 'FECHA DE ENVIO', 'ORIGEN', 'DESCRIPCION', 'CANTIDAD']]
    df_camiones.columns = ['id_camion', 'fecha_salida', 'origen', 'producto', 'cantidad']
    df_camiones.to_csv('data/camiones.csv', index=False)
    print('camiones.csv generado con éxito')

    # 2. Crear inventario.csv
    # Leer el archivo con múltiples niveles de encabezado
    df_inventario = pd.read_excel(archivo_excel, sheet_name="existencia",header=[0, 1])

    # Seleccionar solo las columnas necesarias
    df_inventario = df_inventario[[
        ('ESPAÑOL', 'DESCRIPCIÓN DEL PRODUCTO'),
        ('ESPAÑOL', 'CUPO'), 
        ('RESERVA', 'TOTAL EN RESERVA'), 
        ('RESERVA', 'UBICADO'), 
        ('RESERVA', 'ASIGNADO'), 
        ('RESERVA', 'PARCIALMENTE ASIGNADO'), 
        ('RESERVA', 'PERDIDO')
    ]]

    # Renombrar las columnas
    df_inventario.columns = ['Producto', 'Cupo', 'Total', 'Ubicado', 'Asignado', 'Parcialmente Asignado', 'Perdido']

    # Reordenar las columnas según lo requerido
    df_inventario = df_inventario[['Producto', 'Cupo', 'Total', 'Ubicado', 'Asignado', 'Parcialmente Asignado', 'Perdido']]

    # Guardar el DataFrame en un archivo CSV
    df_inventario.to_csv('data/inventario.csv', index=False)
    print('inventario.csv generado con éxito')

    # 3. Crear ordenes.csv
    wb = load_workbook(archivo_excel, data_only=False)  # data_only=False para obtener las fórmulas tal cual
    ws = wb['ordenes']  # Seleccionar la hoja 'ordenes'

    # Extraer los datos manualmente, omitiendo la primera fila de encabezados
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    # Crear un DataFrame a partir de los datos, empezando desde la segunda fila
    df_ordenes = pd.DataFrame(data[1:], columns=[cell.value for cell in ws[1]])

    # Función para limpiar valores que tienen formato ="3077005"
    def clean_formula(value):
        if isinstance(value, str) and value.startswith('="') and value.endswith('"'):
            return value.strip('="')
        return value

    # Aplicar la función de limpieza a la columna 'Orden'
    df_ordenes['Orden'] = df_ordenes['Orden'].apply(clean_formula)

    # Seleccionar y renombrar las columnas necesarias
    df_ordenes = df_ordenes[['Orden', 'Descripcion de articulo', 'Texto breve personalizado de detalle 5', 
                            'Cantidad de orden original', 'Cantidad solicitada', 'Cantidad asignada', 'orderdtlstatus']]
    df_ordenes.columns = ['id_orden', 'producto', 'tipo_de_orden', 'original', 'solicitada', 'asignada','status']

    # Guardar el DataFrame en un archivo CSV sin la fila extra de nombres
    df_ordenes.to_csv('data/ordenes.csv', index=False)
    print('ordenes.csv generado con éxito')